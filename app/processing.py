from openpyxl import load_workbook

class ProcessPartA:
    def __init__(self):
        self.data = {
            "drillholes": [],
            "samples": [],
            "extra": [],
        }

    def retrieve_data(self, file_name):
        workbook = load_workbook(filename=file_name)

        # Load DRILLHOLES
        worksheet = workbook["DRILLHOLES"]
        headers = [cell.value for cell in worksheet[1]]
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            self.data["drillholes"].append(dict(zip(headers, row)))

        # Load SAMPLES
        worksheet = workbook["SAMPLES"]
        headers = [cell.value for cell in worksheet[1]]
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            self.data["samples"].append(dict(zip(headers, row)))


        # Analyze everything
        print("Analyzing data...")
        drilled = 0
        for hole in self.data["drillholes"]:
            drilled += hole["Length (m)"]
        print(f"Total drilled length: {drilled} meters")

        total_grade = 0
        count = 0
        for sample in self.data["samples"]:
            if sample["Au"] != 0.0:
                total_grade += sample["Au"]
                count += 1
        average_grade = total_grade / count if count > 0 else 0.0
        print(f"Average Au grade: {average_grade}")
